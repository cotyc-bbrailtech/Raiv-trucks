#!/usr/bin/env python3
"""
Raiv-Tracker CM5 HMI log importer (email/IMAP edition)
=========================================================

This script is the receiving half of the "CM5 pushes, GitHub pulls" logging
pipeline:

    CM5 HMI (Log Manager) -- SMTP, attachment --> dedicated mailbox
    GitHub Actions (this script) -- IMAP --> reads mailbox, imports attachments
    -> appends rows into logs/raiv_tracker_<date>.xlsx -> committed by workflow

WHY EMAIL INSTEAD OF FTP
---------------------------
FTP needs something to be listening for an *inbound* connection, which means
either an always-on server (the VPS dependency this whole project has been
trying to avoid) or fighting FTP's dual-channel behavior through StrideLinx's
NAT/port-forwarding. Email flips the direction: the CM5 makes an *outbound*
SMTP connection (which almost every network allows out, unlike inbound
FTP), and GitHub Actions -- which already runs on a schedule -- just checks
a mailbox each time it wakes up. No server to host, ever.

IMPORTANT -- THIS PARSER'S SHAPE IS A BEST GUESS, VERIFY AGAINST A REAL EXPORT
---------------------------------------------------------------------------
The CM5 Hardware User Manual documents *that* the Log Manager can log
numeric tag data and email/FTP the resulting file, but not the *exact* file
layout that C-more's Log Manager produces (that lives in the C-more
Programming Software's own help files, which weren't available while this
was written). This script currently assumes a fairly generic "wide" CSV
export: a timestamp in the first column, one column per logged tag after
that. Once you've configured the Log Manager and have one real sample
export, send it over so this parser can be adjusted to match exactly --
treat everything below as a working first draft, not a guarantee.

Each row of the wide CSV is "melted" into the same long-format rows used
everywhere else in this project: Timestamp | Truck | Tag | Value | Datatype
| Topic -- so this drops into the same logs/raiv_tracker_<date>.xlsx files
that the standalone script and the GitHub Actions MQTT logger both use.
"""

import csv
import email
import imaplib
import io
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


IMAP_HOST = os.environ.get("IMAP_HOST")
IMAP_PORT = int(os.environ.get("IMAP_PORT", "993"))
IMAP_USERNAME = os.environ.get("IMAP_USERNAME")
IMAP_PASSWORD = os.environ.get("IMAP_PASSWORD")
IMAP_MAILBOX = os.environ.get("IMAP_MAILBOX", "INBOX")
# Only process messages whose subject contains this (set to something the
# CM5's Email Client subject line will actually contain, e.g. "RAIV_2 HMI Log").
SUBJECT_FILTER = os.environ.get("SUBJECT_FILTER", "")
TRUCK_LABEL = os.environ.get("TRUCK_LABEL", "Truck 2")
LOG_DIR = Path(os.environ.get("LOG_DIR", "logs"))

HEADERS = ["Timestamp", "Truck", "Tag", "Value", "Datatype", "Topic"]


def log(msg):
    print(f"[{datetime.now().isoformat(timespec='seconds')}] {msg}", flush=True)


class ExcelLogger:
    """Same daily-rotation behaviour as the other loggers in this repo."""

    def __init__(self, log_dir):
        self.log_dir = log_dir
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.current_date = None
        self.wb = None
        self.ws = None
        self.path = None
        self.rows_added = 0

    def _path_for(self, day):
        return self.log_dir / f"raiv_tracker_{day.isoformat()}.xlsx"

    def _ensure_workbook_for(self, day):
        if day == self.current_date:
            return
        if self.wb is not None:
            self.save()
        path = self._path_for(day)
        if path.exists():
            self.wb = load_workbook(path)
            self.ws = self.wb.active
            log(f"Resuming existing log file: {path} ({self.ws.max_row - 1} rows so far today)")
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Log"
            self.ws.append(HEADERS)
            log(f"Started new log file: {path}")
        self.path = path
        self.current_date = day

    def add_row(self, timestamp, truck, tag, value, datatype, topic):
        self._ensure_workbook_for(timestamp.date())
        self.ws.append([
            timestamp.isoformat(timespec="seconds"),
            truck,
            tag,
            value,
            datatype,
            topic,
        ])
        self.rows_added += 1

    def save(self):
        if self.wb is None:
            return
        tmp_path = self.path.with_suffix(".xlsx.tmp")
        self.wb.save(tmp_path)
        os.replace(tmp_path, self.path)


def parse_timestamp(raw):
    """Best-effort parse of whatever timestamp format the Log Manager exports.
    Extend this list once a real sample is available."""
    raw = raw.strip()
    formats = [
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    log(f"Could not parse timestamp {raw!r}, using current time instead")
    return datetime.now()


def melt_wide_csv(csv_bytes, excel_logger, source_label):
    """Assumes column 0 is a timestamp and every other column is a tag whose
    header is the tag name. One input row -> one output row per tag column."""
    text = csv_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return 0
    header = rows[0]
    tag_names = header[1:]
    added = 0
    for row in rows[1:]:
        if not row or not row[0].strip():
            continue
        timestamp = parse_timestamp(row[0])
        for col_index, tag_name in enumerate(tag_names, start=1):
            if col_index >= len(row):
                continue
            value = row[col_index].strip()
            if value == "":
                continue
            excel_logger.add_row(
                timestamp=timestamp,
                truck=TRUCK_LABEL,
                tag=tag_name.strip(),
                value=value,
                datatype="",
                topic=source_label,
            )
            added += 1
    return added


def iter_csv_attachments(msg):
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename = part.get_filename()
        if not filename:
            continue
        if not filename.lower().endswith(".csv"):
            continue
        payload = part.get_payload(decode=True)
        if payload:
            yield filename, payload


def main():
    if not (IMAP_HOST and IMAP_USERNAME and IMAP_PASSWORD):
        log("IMAP_HOST / IMAP_USERNAME / IMAP_PASSWORD are not all set "
            "(expected from IMAP_* Actions secrets). Exiting without checking mail.")
        sys.exit(0)  # don't fail the workflow run over a config issue

    excel_logger = ExcelLogger(LOG_DIR)

    try:
        conn = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
        conn.login(IMAP_USERNAME, IMAP_PASSWORD)
        conn.select(IMAP_MAILBOX)
    except Exception as exc:  # noqa: BLE001
        log(f"Could not connect/login this run: {exc!r} -- will try again next scheduled run.")
        sys.exit(0)

    criteria = "UNSEEN"
    if SUBJECT_FILTER:
        criteria = f'(UNSEEN SUBJECT "{SUBJECT_FILTER}")'
    status, data = conn.search(None, criteria)
    if status != "OK":
        log(f"IMAP search failed: {status}")
        conn.logout()
        sys.exit(0)

    message_ids = data[0].split()
    log(f"Found {len(message_ids)} unseen message(s) matching filter.")

    total_rows = 0
    processed = 0
    for msg_id in message_ids:
        status, msg_data = conn.fetch(msg_id, "(RFC822)")
        if status != "OK" or not msg_data or not msg_data[0]:
            log(f"Could not fetch message {msg_id!r}, skipping.")
            continue
        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)
        subject = msg.get("Subject", "(no subject)")

        found_attachment = False
        for filename, payload in iter_csv_attachments(msg):
            found_attachment = True
            log(f"Importing attachment '{filename}' from message {subject!r}")
            total_rows += melt_wide_csv(payload, excel_logger, source_label=f"email:{filename}")

        if not found_attachment:
            log(f"Message {subject!r} had no .csv attachment, marking seen anyway.")

        # Mark seen regardless, so a message we can't parse doesn't get
        # retried forever and pile up as "unread" indefinitely.
        conn.store(msg_id, "+FLAGS", "\\Seen")
        processed += 1

    conn.logout()
    excel_logger.save()
    log(f"Done. Processed {processed} message(s), logged {total_rows} row(s) this run "
        f"({'no file changes' if total_rows == 0 else excel_logger.path.name}).")


if __name__ == "__main__":
    main()

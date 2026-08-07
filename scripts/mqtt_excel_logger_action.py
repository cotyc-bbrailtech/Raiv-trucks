#!/usr/bin/env python3
"""
Raiv-Tracker MQTT -> Excel logger (GitHub Actions edition)
===========================================================

This is a *bounded* variant of the standalone `mqtt_excel_logger.py` script,
built to run once per scheduled GitHub Actions invocation instead of forever
in the background:

    connect -> listen for LISTEN_SECONDS -> append any readings to today's
    Excel file in logs/ -> exit

The calling workflow (.github/workflows/mqtt-excel-logger.yml) then commits
and pushes logs/ if anything changed. Same "long format" row layout and same
daily-file-rotation logic as the standalone script, so files produced by
either one drop into the same logs/ folder with the same structure:

    Timestamp | Truck | Tag | Value | Datatype | Topic

WHY A BOUNDED LISTEN WINDOW INSTEAD OF FOREVER
------------------------------------------------
GitHub Actions has no "always-on" job type -- every run has to start, do its
work, and finish. Scheduled workflows are the closest thing GitHub offers to
a recurring background job, so this script listens for a short window each
time it's invoked (default 90 seconds) and captures whatever tag readings
publish during that window, then exits.

COVERAGE TRADE-OFF -- READ THIS
---------------------------------
Tags publish roughly once a minute. If this workflow runs every 10 minutes
and only listens for ~90 seconds each time, it will typically catch ONE
reading per tag per run, not all ~10 readings that occurred in between.
That means roughly 10x fewer data points than the standalone always-on
script would capture over the same span. If you need true minute-by-minute
history, run mqtt_excel_logger.py continuously on an always-on machine
instead (see the main project's README) -- this Actions version trades
resolution for running entirely inside GitHub, at no cost and with no
server to maintain.

GitHub also does not guarantee scheduled workflows fire exactly on time --
they can be delayed, especially during high load, and are disabled
automatically after 60 days of repository inactivity (any commit resets
that clock).

APPROACH 7 -- CM5 NATIVE MQTT CLIENT SUPPORT
-----------------------------------------------
The CM5 HMI has its own built-in MQTT Client/Publisher (Setup ribbon ->
Function tab -> MQTT in C-more Programming Software) -- separate from
StrideLinx's router-side MQTT relay. As actually built on RAIV_2, each
Publisher row uses PayloadType=1 ("raw tag value"), so a message is just
one tag's value as plain text with nothing else in it -- no JSON, no
timestamp, no tag name. Topics follow `raiv-tracker/<truck>/hmi/<group>`,
currently g1..g40 across four Publisher objects of 10 topics each.

Any topic containing "/hmi/" is therefore routed to the CM5 handler rather
than the StrideLinx Sparkplug-B one. Because the payload carries no tag
name, the name is looked up from scripts/hmi_mqtt_tag_map.json, keyed by
the exact topic string. The handler splits the payload on "|" and zips it
positionally against that topic's name list, which means it works
unchanged whether a topic carries one value (today) or several (if a
message is ever rebuilt to pack multiple tags into one payload).

Rows parsed from /hmi/ topics are stamped with the time this run received
the message, since the payload has no time in it.

CHANGE-ONLY LOGGING (/hmi/ topics only)
-----------------------------------------
Measured on the first live run: the CM5 publishes every tag roughly 1.4
times per second, so a single 90-second window captured 4,440 messages
that carried only 39 distinct values -- the same numbers over and over.
Writing all of those would add ~640k near-identical rows per day and
crowd Excel's ~1,048,576-row sheet limit within days.

So the /hmi/ handler now keeps a per-run in-memory record of the last
value written for each (topic, tag) and skips a message whose value is
unchanged. The FIRST message for a given tag in a run is always written,
so every run still produces one baseline reading per tag (~40 rows) plus a
row for any value that actually moves mid-run. Net effect: the log becomes
a ~10-minute-resolution time series with intra-run transitions preserved,
instead of thousands of duplicate rows.

Two deliberate limits. The record is per-process, not persisted between
runs, so a tag that never changes still yields one row per run -- that is
wanted, as it doubles as a heartbeat proving the tag is still publishing.
And this applies ONLY to /hmi/ topics: StrideLinx publishes about once a
minute with real per-metric timestamps and shows no duplication, so that
path is left exactly as it was.

Note this trims rows, not broker traffic -- the CM5 still publishes at the
same rate. Slowing the publish trigger in C-more is the lever for traffic.
"""

import json
import os
import ssl
import sys
import threading
import time
from datetime import datetime
from pathlib import Path

import paho.mqtt.client as mqtt
from openpyxl import Workbook, load_workbook


MQTT_HOST = os.environ.get("MQTT_HOST", "353a1aaf7a42437797d72a6699405ee1.s1.eu.hivemq.cloud")
MQTT_PORT = int(os.environ.get("MQTT_PORT", "8883"))
MQTT_USERNAME = os.environ.get("MQTT_USERNAME", "Raiv-Excel-Logger")
MQTT_PASSWORD = os.environ.get("MQTT_PASSWORD")
TOPIC_FILTER = os.environ.get("TOPIC_FILTER", "raiv-tracker/#")
LOG_DIR = Path(os.environ.get("LOG_DIR", "logs"))
LISTEN_SECONDS = float(os.environ.get("LISTEN_SECONDS", "90"))
HMI_TAG_MAP_PATH = Path(os.environ.get("HMI_TAG_MAP_PATH", "scripts/hmi_mqtt_tag_map.json"))

HEADERS = ["Timestamp", "Truck", "Tag", "Value", "Datatype", "Topic"]


def log(msg):
    print(f"[{datetime.now().isoformat(timespec='seconds')}] {msg}", flush=True)


def truck_label(topic):
    parts = topic.split("/")
    raw_id = parts[1] if len(parts) > 1 else topic
    digits = "".join(ch for ch in raw_id if ch.isdigit())
    return f"Truck {digits}" if digits else raw_id


def load_hmi_tag_map():
    """Loads the topic -> ordered-tag-name-list mapping used to interpret
    CM5-native-MQTT payloads (see the Approach 7 note in this file's
    docstring). Missing/unreadable file just means no /hmi/ topics will be
    recognized yet -- not a fatal error, since this feature is opt-in."""
    if not HMI_TAG_MAP_PATH.exists():
        return {}
    try:
        with open(HMI_TAG_MAP_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("topics", {})
    except (json.JSONDecodeError, OSError) as exc:
        log(f"Could not load HMI tag map from {HMI_TAG_MAP_PATH}: {exc!r} -- "
            f"/hmi/ topics will be skipped this run.")
        return {}


class ExcelLogger:
    """Same daily-rotation behaviour as the standalone script's ExcelLogger,
    trimmed down since this process only ever lives for LISTEN_SECONDS."""

    def __init__(self, log_dir):
        self.log_dir = log_dir
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.current_date = None
        self.wb = None
        self.ws = None
        self.path = None
        self.rows_added = 0

    def _path_for(self, day):
        return self.log_dir / f"raiv_tracker_{day.isoformat()}.xlsx"

    def _ensure_workbook_for(self, day):
        if day == self.current_date:
            return
        if self.wb is not None:
            self.save()
        path = self._path_for(day)
        if path.exists():
            self.wb = load_workbook(path)
            self.ws = self.wb.active
            log(f"Resuming existing log file: {path} ({self.ws.max_row - 1} rows so far today)")
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Log"
            self.ws.append(HEADERS)
            log(f"Started new log file: {path}")
        self.path = path
        self.current_date = day

    def add_row(self, timestamp, truck, tag, value, datatype, topic):
        self._ensure_workbook_for(timestamp.date())
        self.ws.append([
            timestamp.isoformat(timespec="seconds"),
            truck,
            tag,
            value,
            datatype,
            topic,
        ])
        self.rows_added += 1

    def save(self):
        if self.wb is None:
            return
        tmp_path = self.path.with_suffix(".xlsx.tmp")
        self.wb.save(tmp_path)
        os.replace(tmp_path, self.path)


def main():
    if not MQTT_PASSWORD:
        log("MQTT_PASSWORD is not set (expected from the HIVEMQ_MQTT_PASSWORD "
            "Actions secret). Exiting without logging anything this run.")
        sys.exit(0)  # don't fail the workflow run over a config issue

    excel_logger = ExcelLogger(LOG_DIR)
    received_lock = threading.Lock()
    hmi_tag_map = load_hmi_tag_map()
    if hmi_tag_map:
        log(f"Loaded HMI tag map with {len(hmi_tag_map)} topic(s) from {HMI_TAG_MAP_PATH}")

    # Last value written this run, keyed by (topic, tag) -- see the
    # "CHANGE-ONLY LOGGING" note in this file's docstring. Only /hmi/ rows
    # consult this; the StrideLinx path is unaffected.
    hmi_last_value = {}
    hmi_skipped = 0

    def on_connect(client, userdata, connect_flags, reason_code, properties):
        if reason_code == 0:
            log(f"Connected to {MQTT_HOST}:{MQTT_PORT}")
            client.subscribe(TOPIC_FILTER, qos=0)
            log(f"Subscribed to '{TOPIC_FILTER}', listening for {LISTEN_SECONDS:.0f}s...")
        else:
            log(f"Connect failed: {reason_code}")

    def handle_hmi_message(msg):
        """CM5 native-MQTT-client payload: one raw tag value (or 'v1|v2|...'
        if a topic is ever built to carry several) with no tag names or
        timestamp in the message itself -- see load_hmi_tag_map() and the
        Approach 7 docstring note above.

        Writes a row only when the value differs from the last one written
        for that (topic, tag) this run; see "CHANGE-ONLY LOGGING" above for
        why, and note the first message per tag always gets written."""
        nonlocal hmi_skipped
        tag_names = hmi_tag_map.get(msg.topic)
        if not tag_names:
            log(f"No tag-map entry for HMI topic '{msg.topic}' -- add one to "
                f"{HMI_TAG_MAP_PATH} (skipping this message).")
            return
        try:
            text = msg.payload.decode("utf-8", errors="replace")
        except Exception:  # noqa: BLE001
            return
        values = text.split("|")
        truck = truck_label(msg.topic)
        now = datetime.now()
        with received_lock:
            for name, value in zip(tag_names, values):
                value = value.strip()
                if value == "":
                    continue
                key = (msg.topic, name)
                if hmi_last_value.get(key) == value:
                    hmi_skipped += 1
                    continue
                hmi_last_value[key] = value
                excel_logger.add_row(
                    timestamp=now,
                    truck=truck,
                    tag=name,
                    value=value,
                    datatype="",
                    topic=msg.topic,
                )

    def on_message(client, userdata, msg):
        if "/hmi/" in msg.topic:
            handle_hmi_message(msg)
            return

        try:
            payload = json.loads(msg.payload.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            log(f"Skipping non-JSON message on {msg.topic}")
            return

        truck = truck_label(msg.topic)
        metrics = payload.get("metrics") if isinstance(payload, dict) else None
        if not metrics:
            metrics = [payload] if isinstance(payload, dict) else []

        with received_lock:
            for metric in metrics:
                name = metric.get("name")
                if not name:
                    continue
                timestamp = metric_timestamp(metric)
                excel_logger.add_row(
                    timestamp=timestamp,
                    truck=truck,
                    tag=name,
                    value=metric.get("value"),
                    datatype=metric.get("datatype"),
                    topic=msg.topic,
                )

    def metric_timestamp(metric):
        ts = metric.get("timestamp")
        if isinstance(ts, (int, float)):
            try:
                seconds = ts / 1000 if ts > 1e12 else ts
                return datetime.fromtimestamp(seconds)
            except (ValueError, OSError, OverflowError):
                pass
        return datetime.now()

    client = mqtt.Client(
        mqtt.CallbackAPIVersion.VERSION2,
        client_id=f"raiv-excel-logger-action-{os.getpid()}",
        clean_session=True,
    )
    client.username_pw_set(MQTT_USERNAME, MQTT_PASSWORD)
    client.tls_set(cert_reqs=ssl.CERT_REQUIRED)
    client.on_connect = on_connect
    client.on_message = on_message

    try:
        client.connect(MQTT_HOST, MQTT_PORT, keepalive=60)
    except Exception as exc:  # noqa: BLE001
        log(f"Could not connect this run: {exc!r} -- will try again next scheduled run.")
        sys.exit(0)

    client.loop_start()
    time.sleep(LISTEN_SECONDS)
    client.loop_stop()
    client.disconnect()

    excel_logger.save()
    if hmi_skipped:
        log(f"Skipped {hmi_skipped} unchanged /hmi/ message(s) "
            f"({len(hmi_last_value)} distinct tag(s) seen) -- change-only logging.")
    log(f"Done. Logged {excel_logger.rows_added} rows this run "
        f"({'no file changes' if excel_logger.rows_added == 0 else excel_logger.path.name}).")


if __name__ == "__main__":
    main()
